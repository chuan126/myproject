"""测试循环测试设备 Excel 工作簿的解析与转换功能。

验证以下核心行为：
1. 从包含 record、auxAdapter、auxTemp 等工作表的 Excel 工作簿正确提取和合并数据列
2. 电流符号按充放电工步类型正确设置（充电为正、放电为负）
3. 容量的积分计算（cc_capacity，单位 Ah）
4. 力学特征（delta_f、df_dt、df_dq、force_slope）的派生计算
5. 批量转换的进度条显示、manifest 文件生成
6. 覆盖写入时正确清理旧的序列文件
"""

import tempfile
import unittest
from pathlib import Path
from unittest.mock import MagicMock, patch

import pandas as pd
import yaml

from src.data.converters.cycler_workbook import (
    convert_cycler_workbook,
    _force_column,
    prepare_cycler_workbook,
    prepare_cycler_workbooks,
)
from src.data.converters.features import derive_mechanical_features


def _sequence(sequence_id: str) -> dict[str, pd.DataFrame]:
    """构造单序列的模拟 DataFrame，用于 mock 测试。

    包含 time、soc、sequence_id、voltage 四列，共两行数据。

    Args:
        sequence_id: 序列标识符。

    Returns:
        以 sequence_id 为键、DataFrame 为值的字典。
    """
    return {
        sequence_id: pd.DataFrame(
            {
                "time": [0.0, 1.0],
                "soc": [0.0, 1.0],
                "sequence_id": [sequence_id, sequence_id],
                "voltage": [3.0, 4.2],
            }
        )
    }


class PrepareCyclerWorkbooksTests(unittest.TestCase):
    """测试 Excel 工作簿到规范化 CSV 的完整转换流程。"""

    def test_converts_workbook_from_record_aux_adapter_and_t3(self) -> None:
        """验证从包含 record、auxAdapter、auxTemp 三张工作表的 Excel 中正确提取并合并数据。

        构造一个包含充放电工步的模拟工作簿，验证：
        - 输出列集合与预期一致（id, time, voltage, current, power 等）
        - 时间列从绝对时间转换为相对秒数（从 0 开始）
        - 电流符号：充电为正、放电为负
        - 力值从 auxAdapter 的 PV1 列提取
        - 温度从 auxTemp 的 T3 列提取
        - cc_capacity 正确按充放电方向累积（单位 Ah = As/3600）
        - sequence_id 列为工作簿文件名（不含扩展名）
        """
        try:
            from openpyxl import Workbook
        except ImportError:
            self.skipTest("openpyxl is required for workbook conversion tests")

        with tempfile.TemporaryDirectory() as temp_dir:
            workbook_path = Path(temp_dir) / "2000N_25degC_0.5C.xlsx"
            workbook = Workbook()
            # 创建 record 工作表：记录充放电工步数据
            record = workbook.active
            record.title = "record"
            record.append(["数据序号", "工步类型", "绝对时间", "电流(A)", "电压(V)"])
            record.append([10, "恒流充电", "2026-05-29 23:32:14.000", 1.0, 3.0])
            record.append([11, "恒流充电", "2026-05-29 23:32:15.000", 1.0, 3.1])
            record.append([12, "恒流放电", "2026-05-29 23:32:16.000", 1.0, 3.2])
            record.append([13, "恒流放电", "2026-05-29 23:32:17.000", 1.0, 3.3])

            # 创建 auxAdapter 工作表：包含力传感器数据 (PV1)
            aux_adapter = workbook.create_sheet("auxAdapter")
            aux_adapter.append([])
            aux_adapter.append(["数据序号", "绝对时间", "PV1"])
            aux_adapter.append([10, "2026-05-29 23:32:14.000", 2000.0])
            aux_adapter.append([11, "2026-05-29 23:32:15.000", 2001.0])
            aux_adapter.append([12, "2026-05-29 23:32:16.000", 2003.0])
            aux_adapter.append([13, "2026-05-29 23:32:17.000", 2006.0])

            # 创建 auxTemp 工作表：包含温度数据 (T1, T3)
            aux_temp = workbook.create_sheet("auxTemp")
            aux_temp.append([])
            aux_temp.append(["数据序号", "绝对时间", "T1", "T3"])
            aux_temp.append([10, "2026-05-29 23:32:14.000", 99.0, 25.0])
            aux_temp.append([11, "2026-05-29 23:32:15.000", 99.0, 25.1])
            aux_temp.append([12, "2026-05-29 23:32:16.000", 99.0, 25.2])
            aux_temp.append([13, "2026-05-29 23:32:17.000", 99.0, 25.3])
            workbook.save(workbook_path)

            sequences = convert_cycler_workbook(workbook_path)

        frame = sequences["2000N_25degC_0.5C"]
        # 验证输出包含所有期望的列
        self.assertEqual(
            frame.columns.tolist(),
            [
                "id",
                "time",
                "voltage",
                "current",
                "power",
                "cc_capacity",
                "force",
                "temperature",
                "delta_f",
                "delta_q",
                "df_dt",
                "df_dq",
                "force_slope",
                "soc",
                "sequence_id",
            ],
        )
        # 验证 id 列从 1 开始重新编号
        self.assertEqual(frame["id"].tolist(), [1, 2, 3, 4])
        # 验证时间列从 0 开始偏移
        self.assertEqual(frame["time"].tolist(), [0.0, 1.0, 2.0, 3.0])
        # 验证力值从 PV1 正确提取
        self.assertEqual(frame["force"].tolist(), [2000.0, 2001.0, 2003.0, 2006.0])
        # 验证温度从 T3 正确提取
        self.assertEqual(frame["temperature"].tolist(), [25.0, 25.1, 25.2, 25.3])
        # 验证电流符号：充电为正，放电为负
        self.assertEqual(frame["current"].tolist(), [1.0, 1.0, -1.0, -1.0])
        # 验证内部列不出现在输出中
        self.assertNotIn("record_id", frame.columns)
        self.assertNotIn("cycle_id", frame.columns)
        self.assertNotIn("source", frame.columns)
        # 验证容量积分：1A * 1s / 3600 = 1/3600 Ah
        self.assertAlmostEqual(frame["cc_capacity"].iloc[1], 1.0 / 3600.0)
        self.assertAlmostEqual(frame["cc_capacity"].iloc[3], -1.0 / 3600.0)
        # 验证序列标识符为工作簿文件名
        self.assertEqual(frame["sequence_id"].unique().tolist(), ["2000N_25degC_0.5C"])

    def test_detects_force_column_names_from_real_workbooks(self) -> None:
        """验证 _force_column 能正确识别不同命名的力传感器列。

        真实工作簿中力列可能有多种命名方式：
        - "Pressure（N）"：旧版命名
        - "F"：新版简化命名
        - 空字符串：力列缺失时返回默认索引 1
        """
        self.assertEqual(_force_column(["T3", "Pressure（N）"]), 1)
        self.assertEqual(_force_column(["T3", "F"]), 1)
        self.assertEqual(_force_column(["T3", ""]), 1)

    def test_derives_mechanical_features(self) -> None:
        """验证力学派生特征的计算正确性。

        构造三行数据，验证：
        - 第一行（基线）：所有 delta 和导数为 0
        - 第二行：delta_f=4, delta_q=0.5, df_dt=2.0, df_dq=8.0, force_slope=8.0
        - 第三行：df_dq=4/1.5（容量变化非零时的除法），force_slope=4.0（容量变化为零时用力的差值）
        """
        rows = [
            {"time": 0.0, "force": 3000.0, "cc_capacity": 0.0},
            {"time": 2.0, "force": 3004.0, "cc_capacity": 0.5},
            {"time": 4.0, "force": 3008.0, "cc_capacity": 2.0},
        ]

        derive_mechanical_features(rows)

        self.assertEqual(rows[0]["delta_f"], 0.0)
        self.assertEqual(rows[0]["delta_q"], 0.0)
        self.assertEqual(rows[0]["df_dt"], 0.0)
        self.assertEqual(rows[0]["df_dq"], 0.0)
        self.assertEqual(rows[0]["force_slope"], 0.0)
        self.assertEqual(rows[1]["delta_f"], 4.0)
        self.assertEqual(rows[1]["delta_q"], 0.5)
        self.assertEqual(rows[1]["df_dt"], 2.0)
        self.assertEqual(rows[1]["df_dq"], 8.0)
        self.assertEqual(rows[1]["force_slope"], 8.0)
        self.assertAlmostEqual(rows[2]["df_dq"], 4.0 / 1.5)
        self.assertEqual(rows[2]["force_slope"], 4.0)

    @patch("src.data.converters.cycler_workbook.tqdm")
    @patch("src.data.converters.cycler_workbook.convert_cycler_workbook")
    def test_shows_tqdm_progress_by_default(self, convert_mock, tqdm_mock) -> None:
        """验证批量转换时默认显示 tqdm 进度条。

        Mock 掉实际转换逻辑，只验证：
        - tqdm 被调用且 disable=False（默认启用）
        - 进度条描述为 "Converting Excel workbooks"
        - 进度条单位设置为 "file"
        - 进度条 postfix 显示当前正在处理的工作簿文件名
        """
        convert_mock.return_value = _sequence("a_cycle_1")
        progress = MagicMock()
        tqdm_mock.return_value = progress

        with tempfile.TemporaryDirectory() as temp_dir:
            workbook_path = Path(temp_dir) / "a.xlsx"
            workbook_path.write_bytes(b"a")
            progress.__iter__.return_value = iter([workbook_path])
            prepare_cycler_workbooks([workbook_path], Path(temp_dir), overwrite=True)

        tqdm_mock.assert_called_once_with(
            [workbook_path],
            desc="Converting Excel workbooks",
            unit="file",
            disable=False,
        )
        progress.set_postfix_str.assert_called_once_with("a.xlsx")

    @patch("src.data.converters.cycler_workbook.convert_cycler_workbook")
    def test_writes_one_manifest_for_multiple_workbooks(self, convert_mock) -> None:
        """验证批量转换多个工作簿时生成一个合并的 manifest 文件。

        manifest.yaml 应包含：
        - dataset_name：输出目录名
        - sequence_count：生成的 CSV 文件数（每个工作簿一个序列）
        - row_count：所有序列的总行数
        - raw_files：原始工作簿文件路径列表
        - raw_file_signatures：每个原始文件的签名信息
        """
        convert_mock.side_effect = [_sequence("a_cycle_1"), _sequence("b_cycle_1")]

        with tempfile.TemporaryDirectory() as temp_dir:
            a_path = Path(temp_dir) / "a.xlsx"
            b_path = Path(temp_dir) / "b.xlsx"
            a_path.write_bytes(b"a")
            b_path.write_bytes(b"b")
            output_dir = Path(temp_dir) / "own_cell"
            generated = prepare_cycler_workbooks(
                [a_path, b_path],
                output_dir,
                overwrite=True,
                show_progress=False,
            )
            with (output_dir / "manifest.yaml").open(encoding="utf-8") as file:
                manifest = yaml.safe_load(file)

        self.assertEqual(len(generated), 2)
        self.assertEqual(manifest["dataset_name"], "own_cell")
        self.assertEqual(manifest["sequence_count"], 2)
        self.assertEqual(manifest["row_count"], 4)
        self.assertEqual(manifest["raw_files"], [str(a_path), str(b_path)])
        self.assertEqual(len(manifest["raw_file_signatures"]), 2)
        self.assertEqual(manifest["raw_file_signatures"][0]["path"], str(a_path.resolve()))

    @patch("src.data.converters.cycler_workbook.convert_cycler_workbook")
    def test_overwrite_removes_stale_sequence_csv(self, convert_mock) -> None:
        """验证覆盖写入时清理不在本次转换结果中的旧 CSV 文件。

        先创建了一个已不存在的序列的 CSV 文件（removed_cycle_1.csv），
        然后执行覆盖写入，验证该文件被删除，而新生成的 CSV 文件存在。
        """
        convert_mock.return_value = _sequence("a_cycle_1")

        with tempfile.TemporaryDirectory() as temp_dir:
            workbook_path = Path(temp_dir) / "a.xlsx"
            workbook_path.write_bytes(b"a")
            output_dir = Path(temp_dir) / "own_cell"
            # 创建不属于当前工作簿的旧序列 CSV
            stale_path = output_dir / "sequences" / "removed_cycle_1.csv"
            stale_path.parent.mkdir(parents=True)
            stale_path.write_text("time,soc,sequence_id\n0,0.5,removed_cycle_1\n", encoding="utf-8")

            prepare_cycler_workbooks([workbook_path], output_dir, overwrite=True, show_progress=False)

            self.assertFalse(stale_path.exists())
            self.assertTrue((output_dir / "sequences" / "a_cycle_1.csv").exists())

    @patch("src.data.converters.cycler_workbook.convert_cycler_workbook")
    def test_single_workbook_api_keeps_raw_file_metadata(self, convert_mock) -> None:
        """验证单工作簿 API (prepare_cycler_workbook) 使用 raw_file（单数）而非 raw_files（复数）。

        单个工作簿转换时 manifest 应使用 raw_file 键名，
        且 dataset_name 默认为工作簿文件名（不含扩展名）。
        """
        convert_mock.return_value = _sequence("a_cycle_1")

        with tempfile.TemporaryDirectory() as temp_dir:
            workbook_path = Path(temp_dir) / "a.xlsx"
            workbook_path.write_bytes(b"a")
            output_dir = Path(temp_dir) / "output"
            prepare_cycler_workbook(workbook_path, output_dir, overwrite=True, show_progress=False)
            with (output_dir / "manifest.yaml").open(encoding="utf-8") as file:
                manifest = yaml.safe_load(file)

        self.assertEqual(manifest["dataset_name"], "a")
        self.assertEqual(manifest["raw_file"], str(workbook_path))
        self.assertNotIn("raw_files", manifest)
        self.assertEqual(manifest["raw_file_signatures"][0]["path"], str(workbook_path.resolve()))


if __name__ == "__main__":
    unittest.main()
