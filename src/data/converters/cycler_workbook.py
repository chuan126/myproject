"""电池循环仪 Excel 工作簿转换模块。

将电池循环仪导出的 Excel 工作簿（.xlsx）转换为规范化的 SOC CSV 数据集。
主要转换流程：
1. 解析工作簿中的 record、auxAdapter、auxTemp 等工作表
2. 执行 1 Hz 去重采样（按秒取整）
3. 安时积分（库仑计数）累积充放电容量
4. 基于充放电容量计算 SOC 值
5. 映射压力和温度等辅助测量值
6. 派生力学特征（delta_f, delta_q, df_dt, df_dq, force_slope）

在整个项目中的角色：
- 位于 data/converters 子包，是数据转换层
- 被 scripts 中的准备脚本调用，将原始实验数据转为训练可用格式
- 输出为规范化 CSV + manifest.yaml，可直接被 data 层其他模块消费
"""

from hashlib import sha256
from pathlib import Path
from typing import Any, Iterable

import pandas as pd
import yaml
from tqdm.auto import tqdm

from ..schema import SEQUENCE_COLUMN, SOC_COLUMN, TIME_COLUMN, canonical_manifest
from .features import derive_mechanical_features


def raw_file_signature(path: Path) -> dict[str, Any]:
    """返回可用于识别原始文件内容变化的稳定指纹。

    通过 SHA-256 哈希算法对整个文件内容计算摘要，结合文件大小和绝对路径，
    提供可靠的原始文件版本追踪能力。用于 manifest 中的 source_file_signatures
    字段，便于追溯降采样或转换后数据的原始来源。

    Args:
        path: 原始文件路径

    Returns:
        包含 "path"（绝对路径）、"size"（字节数）、"sha256"（十六进制哈希）的字典
    """
    resolved = Path(path).resolve()
    digest = sha256()
    with resolved.open("rb") as file:
        # 分块读取大文件，每块 1MB，避免一次性加载到内存
        for block in iter(lambda: file.read(1024 * 1024), b""):
            digest.update(block)
    return {
        "path": str(resolved),
        "size": resolved.stat().st_size,
        "sha256": digest.hexdigest(),
    }


def _find_header(sheet: Any, required_tokens: Iterable[str]) -> tuple[int, list[str]]:
    """在工作表中查找包含所有必需标记的表头行。

    逐行扫描工作表，找到第一个包含所有 required_tokens 的行作为表头。
    表头值会被标准化为去除首尾空格的字符串，空单元格映射为空字符串。

    Args:
        sheet: openpyxl 工作表对象
        required_tokens: 必须在表头行中出现的列名标记（如 ["工步类型", "绝对时间"]）

    Returns:
        (行号, 表头字符串列表) 元组，行号从 1 开始

    Raises:
        ValueError: 没有找到包含所有必需标记的表头行时抛出
    """
    if hasattr(sheet, "reset_dimensions"):
        sheet.reset_dimensions()
    for row_number, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        headers = [str(value).strip() if value is not None else "" for value in row]
        if all(token in headers for token in required_tokens):
            return row_number, headers
    raise ValueError(f"Worksheet '{sheet.title}' has no expected header row: {list(required_tokens)}")


def _column(headers: list[str], candidates: Iterable[str]) -> int:
    """在表头列表中查找匹配候选名的列索引。

    按候选名顺序依次尝试匹配，返回第一个匹配的列索引。
    这允许适配不同语言或命名习惯的工作簿列名（如中文 "电压(V)" 和英文 "voltage"）。

    Args:
        headers: 表头字符串列表
        candidates: 候选列名列表（按优先级从高到低排列）

    Returns:
        匹配到的列索引（从 0 开始）

    Raises:
        ValueError: 所有候选名都不在表头中时抛出
    """
    for candidate in candidates:
        if candidate in headers:
            return headers.index(candidate)
    raise ValueError(f"Missing expected workbook column; accepted names: {list(candidates)}")


def _as_float(value: Any, column_name: str) -> float:
    """将单元格值安全转换为浮点数。

    如果值无法转换为 float（如空值、非数字字符串），抛出包含列名和原始值的详细错误。

    Args:
        value: 单元格原始值
        column_name: 列名（用于错误信息）

    Returns:
        转换后的浮点数

    Raises:
        ValueError: 值无法转换为数值时抛出
    """
    try:
        return float(value)
    except (TypeError, ValueError) as error:
        raise ValueError(f"Workbook value in '{column_name}' is not numeric: {value!r}") from error


def _absolute_time_key(value: Any, column_name: str = "绝对时间") -> int:
    """将绝对时间值转换为纳秒级整数时间戳键。

    用于在 auxAdapter 和 auxTemp 工作表中做精确时间匹配。
    使用 pd.Timestamp 解析各种日期时间格式，输出为纳秒级 Unix 时间戳。

    Args:
        value: 绝对时间值（支持 pandas 可解析的各种格式）
        column_name: 列名（用于错误信息）

    Returns:
        纳秒级整数时间戳

    Raises:
        ValueError: 值不是有效的绝对时间格式时抛出
    """
    try:
        timestamp = pd.Timestamp(value)
    except (TypeError, ValueError) as error:
        raise ValueError(f"Workbook value in '{column_name}' is not a valid absolute time: {value!r}") from error
    if pd.isna(timestamp):
        raise ValueError(f"Workbook value in '{column_name}' is not a valid absolute time: {value!r}")
    return int(timestamp.value)


def _elapsed_seconds(value: Any, origin: Any) -> float:
    """计算从 origin 时间到 value 时间的经过秒数。

    两个时间值都通过 _absolute_time_key 转换为纳秒时间戳后计算差值，
    再除以 1e9 转换为秒。

    Args:
        value: 目标绝对时间值
        origin: 基准绝对时间值

    Returns:
        经过的秒数（浮点数）
    """
    return (_absolute_time_key(value) - _absolute_time_key(origin)) / 1_000_000_000


def _normalized_current(current: float, step_type: str) -> float:
    """根据工步类型归一化电流符号。

    约定：充电电流为正，放电电流为负。
    - 如果工步类型包含 "充电" 或 "charge"，返回正电流（取绝对值）
    - 如果包含 "放电" 或 "discharge"，返回负电流（取负绝对值）
    - 否则保持原始符号

    Args:
        current: 原始电流值（可能为正或负，取决于仪器设置）
        step_type: 工步类型描述字符串

    Returns:
        符号归一化后的电流值
    """
    if "充电" in step_type or "charge" in step_type.lower():
        return abs(current)
    if "放电" in step_type or "discharge" in step_type.lower():
        return -abs(current)
    return current


def _normalized_header(value: Any) -> str:
    """将表头值标准化为统一格式。

    转换为小写、去除首尾空格、统一全角半角括号。
    用于表头的模糊匹配，处理不同来源工作簿的命名差异。

    Args:
        value: 原始表头值

    Returns:
        标准化后的字符串
    """
    return str(value or "").strip().lower().replace("（", "(").replace("）", ")")


def _force_column(headers: list[str]) -> int | None:
    """在表头中查找压力/力传感器的列索引。

    支持多种常见命名：
    - "force"、"f"、"pressure(n)"、"pressure (n)"
    - 以 "pressure" 开头且包含 "(n" 的列
    - 紧跟在 "t3" 列之后的未命名列（某些工作簿中的隐式压力列）

    Args:
        headers: 表头字符串列表

    Returns:
        压力列的索引（从 0 开始），如果找不到则返回 None
    """
    candidates = {"force", "f", "pressure(n)", "pressure (n)"}
    for index, header in enumerate(headers):
        normalized = _normalized_header(header)
        if normalized in candidates:
            return index
        if normalized.startswith("pressure") and "(n" in normalized:
            return index
    # 某些工作簿中将压力列紧跟在 "t3" 列之后，且该列没有明确的表头
    for index, header in enumerate(headers[:-1]):
        if _normalized_header(header) == "t3" and not _normalized_header(headers[index + 1]):
            return index + 1
    return None


def _read_aux_time_map(workbook: Any, sheet_name: str, value_column: str) -> dict[int, float]:
    """从辅助工作表中读取时间到测量值的映射。

    辅助工作表（如 auxAdapter、auxTemp）以绝对时间为键记录压力/温度等测量值。
    此函数将它们读取为以纳秒时间戳为键的字典，供主 record 表查找。

    Args:
        workbook: openpyxl 工作簿对象
        sheet_name: 辅助工作表名称（如 "auxAdapter"、"auxTemp"）
        value_column: 值列名（如 "PV1"、"T3"）

    Returns:
        以纳秒整数时间戳为键、测量值为值的字典
    """
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"Cycler workbook must contain a {sheet_name} worksheet.")
    sheet = workbook[sheet_name]
    header_row, headers = _find_header(sheet, ["绝对时间", value_column])
    time_index = _column(headers, ["绝对时间", "absolute_time"])
    value_index = _column(headers, [value_column])
    values: dict[int, float] = {}
    for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
        if row[time_index] is None:
            continue
        values[_absolute_time_key(row[time_index], headers[time_index])] = _as_float(
            row[value_index],
            headers[value_index],
        )
    return values


def convert_cycler_workbook(
    workbook_path: Path,
    extra_record_columns: Iterable[str] = (),
) -> dict[str, pd.DataFrame]:
    """从导出的循环仪工作簿中提取每个循环的规范化数据帧。

    这是转换的核心函数。对一个工作簿执行完整的转换流水线：

    1. 加载 Excel 工作簿（使用 openpyxl，只读模式，仅读取值）
    2. 从 auxAdapter 工作表读取压力映射（PV1 列）
    3. 从 auxTemp 工作表读取温度映射（T3 列）
    4. 从 record 工作表逐行解析：
       - 按秒取整实现 1 Hz 去重采样（同一秒内只保留第一条记录）
       - 安时积分计算充放电累计容量（charge_ah, discharge_ah）
       - 映射压力和温度测量值
       - 提取电压、电流、功率和可选额外列
    5. 调用 derive_mechanical_features 派生力学特征
    6. 基于充放电容量计算 SOC：
       - 充电工步：SOC = 充电安时 / 总充电安时
       - 放电工步：SOC = 1 - 放电安时 / 总放电安时
    7. 清理中间列（step_type, _charge_ah, _discharge_ah）

    Args:
        workbook_path: 循环仪导出的 Excel 工作簿路径
        extra_record_columns: 需要额外保留的 record 表列名（如实验条件列）

    Returns:
        键为序列名（工作簿文件名不含扩展名）、值为规范化 DataFrame 的字典

    Raises:
        ImportError: 未安装 openpyxl 可选依赖时抛出
        ValueError: 工作簿缺少必需的 worksheet 或数据行时抛出
    """
    try:
        from openpyxl import load_workbook
    except ImportError as error:
        raise ImportError("Converting cycler workbooks requires the optional dependency openpyxl.") from error

    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    if "record" not in workbook.sheetnames:
        raise ValueError("Cycler workbook must contain a record worksheet.")
    record = workbook["record"]
    # 从辅助工作表读取压力和温度的时间序列映射
    force_map = _read_aux_time_map(workbook, "auxAdapter", "PV1")
    temperature_map = _read_aux_time_map(workbook, "auxTemp", "T3")
    header_row, headers = _find_header(record, ["工步类型", "绝对时间"])
    step_type_index = _column(headers, ["工步类型", "step_type"])
    absolute_time_index = _column(headers, ["绝对时间", "absolute_time"])
    current_index = _column(headers, ["电流(A)", "current"])
    voltage_index = _column(headers, ["电压(V)", "voltage"])
    extra_indices = {column: _column(headers, [column]) for column in extra_record_columns}

    rows: list[dict[str, Any]] = []
    sampled_seconds: set[int] = set()
    totals = {"charge": 0.0, "discharge": 0.0}
    last_absolute_time: Any | None = None
    origin_absolute_time: Any | None = None

    for row in record.iter_rows(min_row=header_row + 1, values_only=True):
        if row[absolute_time_index] is None:
            continue
        absolute_time = row[absolute_time_index]
        origin_absolute_time = origin_absolute_time or absolute_time
        sequence_time = _elapsed_seconds(absolute_time, origin_absolute_time)
        step_type = str(row[step_type_index] or "")
        current = _normalized_current(_as_float(row[current_index], headers[current_index]), step_type)
        voltage = _as_float(row[voltage_index], headers[voltage_index])
        previous = last_absolute_time or absolute_time
        # 计算与上一条记录的时间间隔（小时），用于安时积分
        delta_hours = max(0.0, _elapsed_seconds(absolute_time, previous)) / 3600.0
        last_absolute_time = absolute_time
        # 安时积分：累积充放电容量
        totals["charge"] += max(current, 0.0) * delta_hours
        totals["discharge"] += max(-current, 0.0) * delta_hours
        # 1 Hz 去重：同一秒内只保留第一条记录
        rounded_second = round(sequence_time)
        if rounded_second in sampled_seconds:
            continue
        sampled_seconds.add(rounded_second)
        time_key = _absolute_time_key(absolute_time, headers[absolute_time_index])
        if time_key not in force_map:
            raise ValueError(f"auxAdapter worksheet is missing PV1 for absolute time: {absolute_time!r}")
        if time_key not in temperature_map:
            raise ValueError(f"auxTemp worksheet is missing T3 for absolute time: {absolute_time!r}")
        values: dict[str, Any] = {
            TIME_COLUMN: float(rounded_second),
            "voltage": voltage,
            "current": current,
            "power": voltage * current,
            "cc_capacity": totals["charge"] - totals["discharge"],
            "force": force_map[time_key],
            "temperature": temperature_map[time_key],
            "step_type": step_type,
            # 以下划线开头的列为临时中间列，将在 SOC 计算后被删除
            "_charge_ah": totals["charge"],
            "_discharge_ah": totals["discharge"],
        }
        for column, index in extra_indices.items():
            values[column] = row[index]
        rows.append(values)

    derive_mechanical_features(rows)
    charge_total = totals["charge"]
    discharge_total = totals["discharge"]
    if charge_total <= 0.0 or discharge_total <= 0.0:
        raise ValueError(f"Workbook {workbook_path} must contain both charging and discharging records to derive SOC.")
    # 基于充放电容量计算每个时间步的 SOC
    for row_id, values in enumerate(rows, start=1):
        step_type = str(values["step_type"])
        if "充电" in step_type or "charge" in step_type.lower():
            soc = values["_charge_ah"] / charge_total
        elif "放电" in step_type or "discharge" in step_type.lower():
            soc = 1.0 - values["_discharge_ah"] / discharge_total
        elif values["_charge_ah"] <= 0.0:
            soc = 0.0
        elif values["_discharge_ah"] <= 0.0:
            soc = 1.0
        else:
            soc = 1.0 - values["_discharge_ah"] / discharge_total
        values["id"] = row_id
        # SOC 裁剪到 [0, 1] 范围
        values[SOC_COLUMN] = min(1.0, max(0.0, float(soc)))
        values[SEQUENCE_COLUMN] = workbook_path.stem
        # 删除中间计算列，保持输出整洁
        del values["step_type"]
        del values["_charge_ah"]
        del values["_discharge_ah"]
    workbook.close()
    if not rows:
        raise ValueError(f"Cycler workbook contains no usable record rows: {workbook_path}")
    frame = pd.DataFrame(rows)
    # 定义标准列顺序，便于输出一致性
    ordered_columns = [
        "id",
        TIME_COLUMN,
        "voltage",
        "current",
        "power",
        "cc_capacity",
        "force",
        "temperature",
        "delta_f",
        "delta_q",
        "df_dt",
        "df_dq",
        "force_slope",
        SOC_COLUMN,
        SEQUENCE_COLUMN,
    ]
    extra_columns = [column for column in frame.columns if column not in ordered_columns]
    return {workbook_path.stem: frame[ordered_columns + extra_columns]}


def prepare_cycler_workbook(
    workbook_path: Path,
    output_dir: Path,
    overwrite: bool = False,
    extra_record_columns: Iterable[str] = (),
    show_progress: bool = True,
) -> list[Path]:
    """将单个循环仪工作簿转换为规范化数据集（单文件便捷接口）。

    实际委托给 prepare_cycler_workbooks 处理。

    Args:
        workbook_path: 循环仪导出的 Excel 工作簿路径
        output_dir: 输出数据集目录
        overwrite: 是否覆盖已有文件
        extra_record_columns: 需要额外保留的列名
        show_progress: 是否显示进度条

    Returns:
        生成的 CSV 文件路径列表
    """
    return prepare_cycler_workbooks(
        [workbook_path],
        output_dir,
        overwrite=overwrite,
        extra_record_columns=extra_record_columns,
        show_progress=show_progress,
    )


def prepare_cycler_workbooks(
    workbook_paths: Iterable[Path],
    output_dir: Path,
    overwrite: bool = False,
    extra_record_columns: Iterable[str] = (),
    show_progress: bool = True,
) -> list[Path]:
    """将多个循环仪工作簿合并转换为一个规范化数据集。

    流程：
    1. 对每个工作簿调用 convert_cycler_workbook 进行转换
    2. 每个序列（循环）写入独立的 CSV 文件到 output_dir/sequences/
    3. 合并所有数据生成统一的 manifest.yaml

    去重机制：使用 dict.fromkeys 去重输入路径，检测不同工作簿是否会造成
    输出文件冲突。

    Args:
        workbook_paths: 循环仪工作簿路径的可迭代对象
        output_dir: 输出数据集目录
        overwrite: 是否覆盖已有文件；为 True 时还会清理不属于当前转换的旧文件
        extra_record_columns: 需要额外保留的列名
        show_progress: 是否通过 tqdm 显示转换进度条

    Returns:
        生成的 CSV 文件路径列表

    Raises:
        ValueError: 没有提供任何工作簿路径时抛出
        ValueError: 多个工作簿解析到相同序列输出文件时抛出
    """
    # 去重保持输入顺序
    paths = list(dict.fromkeys(Path(path) for path in workbook_paths))
    if not paths:
        raise ValueError("At least one cycler workbook path is required.")

    frames: list[pd.DataFrame] = []
    outputs: list[tuple[Path, pd.DataFrame]] = []
    output_paths: set[Path] = set()
    generated: list[Path] = []
    progress = tqdm(paths, desc="Converting Excel workbooks", unit="file", disable=not show_progress)
    for workbook_path in progress:
        progress.set_postfix_str(workbook_path.name)
        sequences = convert_cycler_workbook(workbook_path, extra_record_columns=extra_record_columns)
        for sequence_name, frame in sequences.items():
            output_path = output_dir / "sequences" / f"{sequence_name}.csv"
            if output_path in output_paths:
                raise ValueError(f"Multiple workbooks resolve to the same sequence output file: {output_path}")
            output_paths.add(output_path)
            outputs.append((output_path, frame))
            frames.append(frame)

    # 覆盖模式：清理输出目录中不属于当前转换的旧 CSV 文件
    if overwrite:
        sequence_dir = output_dir / "sequences"
        for existing_path in sequence_dir.rglob("*.csv") if sequence_dir.exists() else ():
            if existing_path not in output_paths:
                existing_path.unlink()

    for output_path, frame in outputs:
        if overwrite or not output_path.exists():
            output_path.parent.mkdir(parents=True, exist_ok=True)
            frame.to_csv(output_path, index=False)
            generated.append(output_path)

    combined = pd.concat(frames, ignore_index=True)
    metadata: dict[str, Any]
    if len(paths) == 1:
        metadata = {"raw_file": str(paths[0])}
    else:
        metadata = {"raw_files": [str(path) for path in paths]}
    metadata["raw_file_signatures"] = [raw_file_signature(path) for path in paths]
    manifest = canonical_manifest(
        dataset_name=paths[0].stem if len(paths) == 1 else output_dir.name,
        source_type="cycler_workbook",
        frame=combined,
        sampling_period_s=1.0,
        soc_method="cycle_charge_discharge_coulomb_counting",
        current_sign="charge_positive",
        **metadata,
    )
    output_dir.mkdir(parents=True, exist_ok=True)
    with (output_dir / "manifest.yaml").open("w", encoding="utf-8") as file:
        yaml.safe_dump(manifest, file, sort_keys=False, allow_unicode=True)
    return generated
